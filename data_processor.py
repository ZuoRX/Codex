"""
数据处理和清洗模块
- 去重
- 数据标准化
- 导出Excel
"""
import os
import json
import glob
import re
from datetime import datetime
from typing import List, Dict, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from loguru import logger

from config import OUTPUT_CONFIG, DATA_FIELDS
from utils.helpers import parse_salary, clean_text, parse_experience, parse_education


class DataProcessor:
    """数据处理器"""

    def __init__(self):
        self.df: Optional[pd.DataFrame] = None
        self.raw_data: List[Dict] = []

    def load_raw_data(self, data_dir: str = None) -> int:
        """从JSON文件加载原始数据"""
        data_dir = data_dir or OUTPUT_CONFIG["raw_data_dir"]
        all_data = []

        json_files = glob.glob(os.path.join(data_dir, "*.json"))
        logger.info(f"找到 {len(json_files)} 个原始数据文件")

        for json_file in json_files:
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    all_data.extend(data)
                logger.debug(f"加载 {json_file}: {len(data)} 条")
            except Exception as e:
                logger.error(f"加载 {json_file} 失败: {e}")

        self.raw_data = all_data
        logger.info(f"共加载 {len(all_data)} 条原始数据")
        return len(all_data)

    def load_data_from_list(self, data: List[Dict]) -> int:
        """直接从列表加载数据"""
        self.raw_data = data
        return len(data)

    def process(self) -> pd.DataFrame:
        """执行完整数据处理流程"""
        logger.info("开始数据处理...")

        # 创建DataFrame
        self.df = pd.DataFrame(self.raw_data)

        if self.df.empty:
            logger.warning("数据为空！")
            return self.df

        logger.info(f"原始数据: {len(self.df)} 条")

        # 确保所有必要字段存在
        for field in DATA_FIELDS:
            if field not in self.df.columns:
                self.df[field] = ""

        # 选择需要的字段
        existing_fields = [f for f in DATA_FIELDS if f in self.df.columns]
        self.df = self.df[existing_fields]

        # 数据清洗步骤
        self._clean_titles()
        self._clean_salaries()
        self._clean_locations()
        self._clean_company_info()
        self._clean_requirements()
        self._clean_descriptions()
        self._remove_duplicates()
        self._filter_valid_records()
        self._add_derived_fields()
        self._sort_data()

        logger.info(f"处理后数据: {len(self.df)} 条")
        return self.df

    def _clean_titles(self):
        """清洗职位名称"""
        self.df["职位名称"] = self.df["职位名称"].apply(
            lambda x: clean_text(str(x)) if pd.notna(x) else ""
        )
        # 过滤明显不相关的职位
        self.df = self.df[self.df["职位名称"].str.len() > 0]

    def _clean_salaries(self):
        """清洗薪资字段"""
        # 重新解析薪资（确保一致性）
        def reparse_salary(row):
            salary_str = str(row.get("薪资范围", "") or "")
            low = row.get("薪资下限_千元")
            high = row.get("薪资上限_千元")

            if pd.isna(low) or pd.isna(high):
                low, high, unit = parse_salary(salary_str)
                return low, high, unit
            return low, high, row.get("薪资单位", "千元/月")

        results = self.df.apply(reparse_salary, axis=1)
        self.df["薪资下限_千元"] = [r[0] for r in results]
        self.df["薪资上限_千元"] = [r[1] for r in results]
        self.df["薪资单位"] = [r[2] for r in results]

        # 计算薪资中位数
        self.df["薪资中位数_千元"] = self.df.apply(
            lambda row: (row["薪资下限_千元"] + row["薪资上限_千元"]) / 2
            if pd.notna(row["薪资下限_千元"]) and pd.notna(row["薪资上限_千元"])
            else None,
            axis=1
        )

    def _clean_locations(self):
        """清洗工作地点"""
        city_mapping = {
            "北京市": "北京", "上海市": "上海", "广州市": "广州",
            "深圳市": "深圳", "杭州市": "杭州", "成都市": "成都",
            "武汉市": "武汉", "南京市": "南京", "西安市": "西安",
            "重庆市": "重庆", "天津市": "天津", "苏州市": "苏州",
            "合肥市": "合肥", "郑州市": "郑州", "长沙市": "长沙",
        }

        def normalize_city(city):
            if pd.isna(city):
                return "未知"
            city = str(city).strip()
            for full, short in city_mapping.items():
                if full in city:
                    return short
            # 只保留城市名（去除省份等前缀）
            city = re.sub(r'[\u4e00-\u9fa5]+省', '', city)
            city = re.sub(r'[\u4e00-\u9fa5]+自治区', '', city)
            return city.strip() or "未知"

        self.df["工作城市"] = self.df["工作城市"].apply(normalize_city)

    def _clean_company_info(self):
        """清洗公司信息"""
        self.df["公司名称"] = self.df["公司名称"].apply(
            lambda x: clean_text(str(x)) if pd.notna(x) else ""
        )
        self.df["公司规模"] = self.df["公司规模"].apply(
            lambda x: str(x).strip() if pd.notna(x) and x else "未知"
        )
        self.df["公司类型"] = self.df["公司类型"].apply(
            lambda x: str(x).strip() if pd.notna(x) and x else "未知"
        )
        self.df["行业类别"] = self.df["行业类别"].apply(
            lambda x: str(x).strip() if pd.notna(x) and x else "未知"
        )

    def _clean_requirements(self):
        """清洗经验和学历要求"""
        self.df["经验要求"] = self.df["经验要求"].apply(
            lambda x: parse_experience(str(x)) if pd.notna(x) and x else "不限"
        )
        self.df["学历要求"] = self.df["学历要求"].apply(
            lambda x: parse_education(str(x)) if pd.notna(x) and x else "不限"
        )

    def _clean_descriptions(self):
        """清洗职位描述"""
        self.df["职位描述"] = self.df["职位描述"].apply(
            lambda x: clean_text(str(x))[:3000] if pd.notna(x) and x else ""
        )
        self.df["福利待遇"] = self.df["福利待遇"].apply(
            lambda x: str(x).strip() if pd.notna(x) and x else ""
        )

    def _remove_duplicates(self):
        """去除重复数据"""
        original_count = len(self.df)

        # 基于职位名称+公司名称+工作城市去重
        self.df = self.df.drop_duplicates(
            subset=["职位名称", "公司名称", "工作城市"],
            keep="first"
        )

        # 去除职位链接相同的记录
        if "职位链接" in self.df.columns:
            has_link = self.df[self.df["职位链接"].notna() & (self.df["职位链接"] != "")]
            no_link = self.df[self.df["职位链接"].isna() | (self.df["职位链接"] == "")]
            has_link = has_link.drop_duplicates(subset=["职位链接"], keep="first")
            self.df = pd.concat([has_link, no_link], ignore_index=True)

        removed = original_count - len(self.df)
        logger.info(f"去重删除 {removed} 条，剩余 {len(self.df)} 条")

    def _filter_valid_records(self):
        """过滤无效记录"""
        original_count = len(self.df)

        # 过滤职位名称为空的记录
        self.df = self.df[self.df["职位名称"].str.len() > 0]

        # 过滤公司名称为空的记录
        self.df = self.df[self.df["公司名称"].str.len() > 0]

        removed = original_count - len(self.df)
        logger.info(f"过滤无效记录 {removed} 条，剩余 {len(self.df)} 条")

    def _add_derived_fields(self):
        """添加衍生字段"""
        # 添加数据月份
        self.df["数据年份"] = datetime.now().year

        # 薪资档次分类
        def salary_level(salary_mid):
            if pd.isna(salary_mid):
                return "面议"
            elif salary_mid < 8:
                return "8K以下"
            elif salary_mid < 12:
                return "8-12K"
            elif salary_mid < 18:
                return "12-18K"
            elif salary_mid < 25:
                return "18-25K"
            elif salary_mid < 35:
                return "25-35K"
            else:
                return "35K以上"

        self.df["薪资档次"] = self.df["薪资中位数_千元"].apply(salary_level)

    def _sort_data(self):
        """数据排序"""
        sort_cols = []
        if "来源网站" in self.df.columns:
            sort_cols.append("来源网站")
        if "工作城市" in self.df.columns:
            sort_cols.append("工作城市")
        if "薪资中位数_千元" in self.df.columns:
            sort_cols.append("薪资中位数_千元")

        if sort_cols:
            self.df = self.df.sort_values(
                sort_cols,
                ascending=[True, True, False],
                na_position="last"
            ).reset_index(drop=True)

    def export_excel(self, output_path: str = None) -> str:
        """导出到Excel（带格式和统计）"""
        if self.df is None or self.df.empty:
            raise ValueError("没有数据可导出，请先运行 process()")

        os.makedirs(OUTPUT_CONFIG["output_dir"], exist_ok=True)
        output_path = output_path or os.path.join(
            OUTPUT_CONFIG["output_dir"],
            OUTPUT_CONFIG["excel_filename"]
        )

        logger.info(f"开始导出Excel到: {output_path}")

        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        # Sheet 1: 主数据表
        self._create_main_sheet(wb)

        # Sheet 2: 薪资分析
        self._create_salary_analysis_sheet(wb)

        # Sheet 3: 城市分布
        self._create_city_analysis_sheet(wb)

        # Sheet 4: 技能需求分析
        self._create_skills_analysis_sheet(wb)

        # Sheet 5: 公司规模分析
        self._create_company_analysis_sheet(wb)

        # Sheet 6: 来源网站统计
        self._create_source_analysis_sheet(wb)

        wb.save(output_path)
        logger.info(f"Excel导出成功: {output_path}，共 {len(self.df)} 条数据")
        return output_path

    def _style_header(self, ws, header_row: int = 1):
        """设置表头样式"""
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    def _style_data_rows(self, ws, start_row: int = 2):
        """设置数据行样式（交替底色）"""
        fill_light = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        align = Alignment(vertical="center", wrap_text=True)

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row)):
            fill = fill_light if row_idx % 2 == 0 else fill_white
            for cell in row:
                cell.fill = fill
                cell.alignment = align

    def _auto_column_width(self, ws, max_width: int = 40):
        """自动调整列宽"""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    # 中文字符宽度约为英文的1.5倍
                    chinese_chars = sum(1 for c in str(cell.value or "") if '\u4e00' <= c <= '\u9fa5')
                    cell_len = cell_len + chinese_chars * 0.5
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            adjusted_width = min(max_len + 2, max_width)
            ws.column_dimensions[col_letter].width = max(adjusted_width, 8)

    def _create_main_sheet(self, wb: Workbook):
        """创建主数据表"""
        ws = wb.create_sheet("招聘数据_全量", 0)

        # 列名映射（中文显示名 -> DataFrame列名）
        columns = [
            ("序号", None),
            ("职位名称", "职位名称"),
            ("薪资范围", "薪资范围"),
            ("薪资下限(K)", "薪资下限_千元"),
            ("薪资上限(K)", "薪资上限_千元"),
            ("薪资中位数(K)", "薪资中位数_千元"),
            ("薪资档次", "薪资档次"),
            ("工作城市", "工作城市"),
            ("工作区域", "工作区域"),
            ("公司名称", "公司名称"),
            ("公司规模", "公司规模"),
            ("公司类型", "公司类型"),
            ("行业类别", "行业类别"),
            ("经验要求", "经验要求"),
            ("学历要求", "学历要求"),
            ("技能要求", "技能要求"),
            ("福利待遇", "福利待遇"),
            ("招聘人数", "招聘人数"),
            ("发布时间", "发布时间"),
            ("来源网站", "来源网站"),
            ("职位链接", "职位链接"),
            ("爬取时间", "爬取时间"),
        ]

        # 写入表头
        headers = [col[0] for col in columns]
        ws.append(headers)
        self._style_header(ws)

        # 冻结首行
        ws.freeze_panes = "A2"

        # 写入数据
        for idx, (_, row) in enumerate(self.df.iterrows(), 1):
            row_data = [idx]
            for _, df_col in columns[1:]:
                if df_col and df_col in self.df.columns:
                    val = row[df_col]
                    if pd.isna(val):
                        val = ""
                    elif isinstance(val, float) and val == int(val):
                        val = int(val)
                    row_data.append(val)
                else:
                    row_data.append("")
            ws.append(row_data)

        # 设置数据行样式
        self._style_data_rows(ws)
        self._auto_column_width(ws)

        # 设置行高
        ws.row_dimensions[1].height = 25
        for row_num in range(2, len(self.df) + 2):
            ws.row_dimensions[row_num].height = 18

        logger.info(f"主数据表创建完成: {len(self.df)} 行")

    def _create_salary_analysis_sheet(self, wb: Workbook):
        """创建薪资分析表"""
        ws = wb.create_sheet("薪资分析", 1)

        ws.append(["数据分析岗位薪资分析报告"])
        ws["A1"].font = Font(size=14, bold=True, color="1F4E79")
        ws.merge_cells("A1:F1")

        ws.append([])
        ws.append(["一、薪资整体统计"])
        ws["A3"].font = Font(bold=True, size=11)

        # 整体薪资统计
        salary_stats = self.df["薪资中位数_千元"].describe()
        stats_header = ["统计项", "值(千元/月)"]
        ws.append(stats_header)
        self._style_header(ws, ws.max_row)

        stats_data = [
            ("有效薪资数量(条)", int(salary_stats.get("count", 0))),
            ("平均薪资中位数", round(float(salary_stats.get("mean", 0)), 1) if salary_stats.get("mean") else "N/A"),
            ("薪资中位数", round(float(salary_stats.get("50%", 0)), 1) if salary_stats.get("50%") else "N/A"),
            ("最低薪资", round(float(salary_stats.get("min", 0)), 1) if salary_stats.get("min") else "N/A"),
            ("最高薪资", round(float(salary_stats.get("max", 0)), 1) if salary_stats.get("max") else "N/A"),
            ("25分位数", round(float(salary_stats.get("25%", 0)), 1) if salary_stats.get("25%") else "N/A"),
            ("75分位数", round(float(salary_stats.get("75%", 0)), 1) if salary_stats.get("75%") else "N/A"),
        ]
        for stat in stats_data:
            ws.append(list(stat))

        ws.append([])
        ws.append(["二、薪资档次分布"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)
        ws.append(["薪资档次", "职位数量", "占比(%)"])
        self._style_header(ws, ws.max_row)

        level_counts = self.df["薪资档次"].value_counts()
        total = len(self.df)
        level_order = ["8K以下", "8-12K", "12-18K", "18-25K", "25-35K", "35K以上", "面议"]
        for level in level_order:
            count = level_counts.get(level, 0)
            pct = round(count / total * 100, 1) if total > 0 else 0
            ws.append([level, count, pct])

        ws.append([])
        ws.append(["三、各城市平均薪资(千元/月)"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)
        ws.append(["城市", "平均薪资", "职位数量", "薪资中位数"])
        self._style_header(ws, ws.max_row)

        city_salary = self.df.groupby("工作城市")["薪资中位数_千元"].agg(["mean", "count", "median"]).round(1)
        city_salary = city_salary[city_salary["count"] >= 5].sort_values("mean", ascending=False)
        for city, row in city_salary.head(20).iterrows():
            ws.append([city, row["mean"], int(row["count"]), row["median"]])

        ws.append([])
        ws.append(["四、经验要求与薪资关系"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)
        ws.append(["经验要求", "平均薪资(K)", "职位数量"])
        self._style_header(ws, ws.max_row)

        exp_salary = self.df.groupby("经验要求")["薪资中位数_千元"].agg(["mean", "count"]).round(1)
        exp_salary = exp_salary.sort_values("mean", ascending=False)
        for exp, row in exp_salary.iterrows():
            ws.append([exp, row["mean"], int(row["count"])])

        self._auto_column_width(ws, max_width=25)
        self._style_data_rows(ws, start_row=5)

    def _create_city_analysis_sheet(self, wb: Workbook):
        """创建城市分布分析表"""
        ws = wb.create_sheet("城市分布", 2)

        ws.append(["数据分析岗位城市分布分析"])
        ws["A1"].font = Font(size=14, bold=True, color="1F4E79")

        ws.append([])
        ws.append(["一、城市职位数量排名"])
        ws["A3"].font = Font(bold=True, size=11)
        ws.append(["排名", "城市", "职位数量", "占比(%)", "平均薪资(K)"])
        self._style_header(ws, 4)

        city_counts = self.df["工作城市"].value_counts()
        total = len(self.df)
        avg_salaries = self.df.groupby("工作城市")["薪资中位数_千元"].mean().round(1)

        for rank, (city, count) in enumerate(city_counts.head(30).items(), 1):
            pct = round(count / total * 100, 1)
            avg_sal = avg_salaries.get(city, "N/A")
            ws.append([rank, city, count, pct, avg_sal])

        ws.append([])
        ws.append(["二、各城市学历要求分布"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)

        # 城市 x 学历的交叉表
        top_cities = city_counts.head(10).index.tolist()
        city_edu = self.df[self.df["工作城市"].isin(top_cities)].groupby(
            ["工作城市", "学历要求"]
        ).size().unstack(fill_value=0)

        header_row = ["城市"] + list(city_edu.columns)
        ws.append(header_row)
        self._style_header(ws, ws.max_row)

        for city, row in city_edu.iterrows():
            ws.append([city] + list(row.values))

        self._auto_column_width(ws)
        self._style_data_rows(ws, start_row=5)

    def _create_skills_analysis_sheet(self, wb: Workbook):
        """创建技能需求分析表"""
        ws = wb.create_sheet("技能需求分析", 3)

        ws.append(["数据分析岗位技能需求分析"])
        ws["A1"].font = Font(size=14, bold=True, color="1F4E79")

        # 统计技能出现频次
        skill_keywords = [
            "Python", "SQL", "Excel", "Tableau", "Power BI", "R语言",
            "SPSS", "SAS", "Hadoop", "Spark", "Hive", "Flink",
            "机器学习", "深度学习", "数据仓库", "ETL", "数据可视化",
            "A/B测试", "用户画像", "MySQL", "PostgreSQL", "MongoDB",
            "Presto", "ClickHouse", "数据建模", "统计分析",
            "Scala", "Kafka", "数据治理", "Java",
        ]

        # 合并技能来源（技能要求 + 职位描述）
        text_combined = (
            self.df["技能要求"].fillna("") + " " +
            self.df["职位描述"].fillna("")
        )

        skill_counts = {}
        for skill in skill_keywords:
            count = text_combined.str.contains(skill, case=False, na=False).sum()
            if count > 0:
                skill_counts[skill] = int(count)

        ws.append([])
        ws.append(["一、技能需求频次统计"])
        ws["A3"].font = Font(bold=True, size=11)
        ws.append(["排名", "技能/工具", "出现次数", "占比(%)"])
        self._style_header(ws, 4)

        total = len(self.df)
        sorted_skills = sorted(skill_counts.items(), key=lambda x: x[1], reverse=True)
        for rank, (skill, count) in enumerate(sorted_skills, 1):
            pct = round(count / total * 100, 1)
            ws.append([rank, skill, count, pct])

        # 按学历要求分类的技能需求
        ws.append([])
        ws.append(["二、不同学历要求的技能需求对比"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)

        for edu in ["本科", "硕士", "博士"]:
            edu_df = self.df[self.df["学历要求"] == edu]
            if len(edu_df) < 10:
                continue
            edu_text = edu_df["技能要求"].fillna("") + " " + edu_df["职位描述"].fillna("")
            ws.append([f"  {edu}学历要求职位的技能统计 (共{len(edu_df)}条)"])
            ws[f"A{ws.max_row}"].font = Font(bold=True)
            ws.append(["技能", "出现次数", "占比(%)"])
            self._style_header(ws, ws.max_row)
            for skill, _ in sorted_skills[:10]:
                count = edu_text.str.contains(skill, case=False, na=False).sum()
                pct = round(count / len(edu_df) * 100, 1)
                ws.append([skill, int(count), pct])

        self._auto_column_width(ws)

    def _create_company_analysis_sheet(self, wb: Workbook):
        """创建公司规模分析表"""
        ws = wb.create_sheet("公司分析", 4)

        ws.append(["数据分析岗位企业分析"])
        ws["A1"].font = Font(size=14, bold=True, color="1F4E79")

        ws.append([])
        ws.append(["一、公司规模分布"])
        ws["A3"].font = Font(bold=True, size=11)
        ws.append(["公司规模", "职位数量", "占比(%)", "平均薪资(K)"])
        self._style_header(ws, 4)

        size_counts = self.df["公司规模"].value_counts()
        total = len(self.df)
        size_salary = self.df.groupby("公司规模")["薪资中位数_千元"].mean().round(1)

        for size, count in size_counts.items():
            pct = round(count / total * 100, 1)
            avg_sal = size_salary.get(size, "N/A")
            ws.append([size, count, pct, avg_sal])

        ws.append([])
        ws.append(["二、行业分布"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)
        ws.append(["行业", "职位数量", "占比(%)", "平均薪资(K)"])
        self._style_header(ws, ws.max_row)

        ind_counts = self.df["行业类别"].value_counts()
        ind_salary = self.df.groupby("行业类别")["薪资中位数_千元"].mean().round(1)

        for industry, count in ind_counts.head(30).items():
            if not industry or industry == "未知":
                continue
            pct = round(count / total * 100, 1)
            avg_sal = ind_salary.get(industry, "N/A")
            ws.append([industry, count, pct, avg_sal])

        ws.append([])
        ws.append(["三、招聘数量最多的公司Top 30"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)
        ws.append(["公司名称", "职位数量", "平均薪资(K)", "所在城市", "行业"])
        self._style_header(ws, ws.max_row)

        company_stats = self.df.groupby("公司名称").agg(
            职位数量=("职位名称", "count"),
            平均薪资=("薪资中位数_千元", "mean"),
            城市=("工作城市", lambda x: x.mode()[0] if not x.empty else ""),
            行业=("行业类别", lambda x: x.mode()[0] if not x.empty else ""),
        ).round(1).sort_values("职位数量", ascending=False)

        for company, row in company_stats.head(30).iterrows():
            ws.append([company, int(row["职位数量"]), row["平均薪资"], row["城市"], row["行业"]])

        self._auto_column_width(ws)

    def _create_source_analysis_sheet(self, wb: Workbook):
        """创建来源网站统计表"""
        ws = wb.create_sheet("数据来源统计", 5)

        ws.append(["各招聘平台数据统计"])
        ws["A1"].font = Font(size=14, bold=True, color="1F4E79")

        ws.append([])
        ws.append(["一、各平台数据量"])
        ws["A3"].font = Font(bold=True, size=11)
        ws.append(["平台", "数据量", "占比(%)", "平均薪资(K)", "数据完整度(%)"])
        self._style_header(ws, 4)

        source_counts = self.df["来源网站"].value_counts()
        total = len(self.df)
        source_salary = self.df.groupby("来源网站")["薪资中位数_千元"].mean().round(1)

        for source, count in source_counts.items():
            pct = round(count / total * 100, 1)
            avg_sal = source_salary.get(source, "N/A")
            # 计算数据完整度（薪资不为空的比例）
            source_df = self.df[self.df["来源网站"] == source]
            completeness = round(
                source_df["薪资中位数_千元"].notna().sum() / len(source_df) * 100, 1
            )
            ws.append([source, count, pct, avg_sal, completeness])

        ws.append([])
        ws.append(["二、数据摘要"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=11)
        summary_data = [
            ("总数据量", len(self.df)),
            ("有效薪资数量", int(self.df["薪资中位数_千元"].notna().sum())),
            ("覆盖城市数", self.df["工作城市"].nunique()),
            ("覆盖公司数", self.df["公司名称"].nunique()),
            ("数据生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        ws.append(["统计项", "数值"])
        self._style_header(ws, ws.max_row)
        for item in summary_data:
            ws.append(list(item))

        self._auto_column_width(ws)

    def generate_report(self) -> dict:
        """生成数据摘要报告"""
        if self.df is None or self.df.empty:
            return {}

        report = {
            "总数据量": len(self.df),
            "来源网站": self.df["来源网站"].value_counts().to_dict(),
            "城市分布Top10": self.df["工作城市"].value_counts().head(10).to_dict(),
            "平均薪资(K)": round(self.df["薪资中位数_千元"].mean(), 1) if self.df["薪资中位数_千元"].notna().any() else "N/A",
            "学历要求分布": self.df["学历要求"].value_counts().to_dict(),
            "经验要求分布": self.df["经验要求"].value_counts().to_dict(),
            "有效薪资数量": int(self.df["薪资中位数_千元"].notna().sum()),
        }
        return report
